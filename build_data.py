import os, math
import numpy as np
import pandas as pd
from openpyxl import load_workbook

def _pick_range(ranges: dict, candidates, label: str) -> str:
    """
    Return the first non-empty entry from `ranges` matching any of the keys in `candidates`.
    Raise a clear error if none are found.
    """
    for k in candidates:
        v = ranges.get(k)
        if isinstance(v, str) and v.strip():
            return v
    raise KeyError(f"Missing '{label}' in ranges (tried keys: {', '.join(candidates)})")

def build_konus_series(folder, sheet_name, ranges, terrain_lookup):
    """
    Returns dict of borehole data:
    {
      BH: {
        "undist": [...],
        "remould": [...],
        "sensitivity": [...],
        "depths": [...],
        "elevs": [...],
        "Z": terrain_level
      }
    }
    """

    excel_extensions = (".xlsx", ".xls", ".xlsm")
    konus_series = {}

    for filename in os.listdir(folder):
        if not filename.endswith(excel_extensions) or filename.startswith("~$"):
            continue
        bh = os.path.splitext(filename)[0]
        Z = terrain_lookup.get(bh)
        if Z is None:
            print(f"⚠️ No terrain level for {bh}, skipping")
            continue

        path = os.path.join(folder, filename)
        try:
            wb = load_workbook(path, data_only=True)
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ Sheet {sheet_name} not in {filename}, skipping")
                continue
            ws = wb[sheet_name]

            und_raw = [cell[0].value for cell in ws[ranges["konus_undist"]]]
            rem_raw = [cell[0].value for cell in ws[ranges["konus_remould"]]]
            dep_raw = [cell[0].value for cell in ws[ranges["depth"]]]

            depths, elevs, undist, remould, sens = [], [], [], [], []
            for u, r, d in zip(und_raw, rem_raw, dep_raw):
                if d is None:
                    continue
                depths.append(d)
                elevs.append(Z - d)

                cu_val = float(u) if u is not None else np.nan
                cur_val = float(r) if r is not None else np.nan

                undist.append(cu_val if np.isfinite(cu_val) else np.nan)
                remould.append(cur_val if np.isfinite(cur_val) else np.nan)

                if np.isfinite(cu_val) and np.isfinite(cur_val) and cur_val != 0:
                    s_val = cu_val / cur_val
                    sens.append(s_val if s_val > 0 else np.nan)
                else:
                    sens.append(np.nan)

            konus_series[bh] = {
                "undist": undist,
                "remould": remould,
                "sensitivity": sens,
                "depths": depths,
                "elevs": elevs,
                "Z": Z,
            }

        except Exception as e:
            print(f"❌ Error reading {filename}: {e}")

    return konus_series

# --- ENAKS ---------------------------------------------------------------
def build_enaks_series(folder, sheet_name, ranges, terrain_lookup):
    """
    Build a dict per borehole with ENAKS strength (cu) and deformation at break ε_f.

    Expected keys in `ranges` (any alias works):
      - strength (kPa):  'enaks_strength' | 'x_range_enaks_strength' | 'strength'
      - deform  (%):     'enaks_deform'   | 'x_range_enaks_deform'   | 'deform'
      - depth (m):       'enaks_depth'    | 'y_range_enaks_depth'    | 'depth'
    Returns:
      {
        "BH01": {
          "Z": <terrain level>,
          "depths": [..], "elevs": [..],
          "strength":[.. or None ..],
          "deform":  [.. or None ..]
        }, ...
      }
    """
    str_rng = _pick_range(ranges, ["enaks_strength","x_range_enaks_strength","strength"], "enaks strength range")
    def_rng = _pick_range(ranges, ["enaks_deform","x_range_enaks_deform","deform"], "enaks deform range")
    dep_rng = _pick_range(ranges, ["enaks_depth","y_range_enaks_depth","depth"], "enaks depth range")

    excel_ext = (".xlsx", ".xls", ".xlsm")
    out = {}

    for fname in os.listdir(folder):
        if not fname.endswith(excel_ext) or fname.startswith("~$"):
            continue
        path = os.path.join(folder, fname)
        bh = os.path.splitext(fname)[0]
        Z = terrain_lookup.get(bh)
        if Z is None:
            print(f"⚠️ Terrain level not found for {bh}, skipping Enaks.")
            continue

        try:
            wb = load_workbook(path, data_only=True)
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ Sheet '{sheet_name}' not in {fname}, skipping.")
                continue
            ws = wb[sheet_name]

            str_raw = [c[0].value for c in ws[str_rng]]
            def_raw = [c[0].value for c in ws[def_rng]]
            dep_raw = [c[0].value for c in ws[dep_rng]]

            depths, elevs, strength, deform = [], [], [], []
            for cu, df, d in zip(str_raw, def_raw, dep_raw):
                if d is None:
                    continue
                depths.append(d)
                elevs.append(Z - d)
                strength.append(float(cu) if cu is not None else None)
                deform.append(float(df) if df is not None else None)

            out[bh] = {
                "Z": Z,
                "depths": depths,
                "elevs": elevs,
                "strength": strength,
                "deform": deform,
            }
        except Exception as e:
            print(f"❌ Error reading {fname}: {e}")

    return out

def build_wc_series(folder, sheet_name, ranges, terrain_lookup):
    """
    Returns dict of borehole data:
    {
      BH: {
        "Z": <terrain level>,
        "depths": [...],
        "elevs": [...],
        "water content": [...],
      }
    }
    """

    excel_extensions = (".xlsx", ".xls", ".xlsm")
    wc_series = {}

    for filename in os.listdir(folder):
        if not filename.endswith(excel_extensions) or filename.startswith("~$"):
            continue
        bh = os.path.splitext(filename)[0]
        Z = terrain_lookup.get(bh)
        if Z is None:
            print(f"⚠️ No terrain level for {bh}, skipping")
            continue

        path = os.path.join(folder, filename)
        try:
            wb = load_workbook(path, data_only=True)
            if sheet_name not in wb.sheetnames:
                print(f"⚠️ Sheet {sheet_name} not in {filename}, skipping")
                continue
            ws = wb[sheet_name]

            wc_raw = [cell[0].value for cell in ws[ranges["wc"]]]
            dep_raw = [cell[0].value for cell in ws[ranges["wc_depth"]]]

            depths, elevs, wc = [], [], []
            for v,d in zip(wc_raw, dep_raw):
                if d is None:
                    continue
                depths.append(d)
                elevs.append(Z - d)
                wc.append(float(v) if v is not None else None)
                
            wc_series[bh] = {
                "Z": Z,
                "depths": depths,
                "elevs": elevs,
                "water content": wc
            }

        except Exception as e:
            print(f"❌ Error reading {filename}: {e}")

    return wc_series

def export_combined_table(konus_series, enaks_series, outfile_xlsx):
    """
    Export combined borehole data to Excel.

    Columns:
      Borhull | Dybde | Kote | Omrørt skjærstyrke | Uforstyrret skjærstyrke konus |
      Sensitivitet | Skjærstyrke enaks | Bruddtøyning
    """

    rows = []
    all_bhs = set(konus_series.keys()) | set(enaks_series.keys())

    for bh in sorted(all_bhs):
        kdata = konus_series.get(bh, {})
        edata = enaks_series.get(bh, {})

        depths = kdata.get("depths", []) or edata.get("depths", [])
        elevs  = kdata.get("elevs", []) or edata.get("elevs", [])

        for i, d in enumerate(depths):
            row = {
                "Borhull": bh,
                "Dybde": d,
                "Kote": elevs[i] if i < len(elevs) else None,
                "Omrørt skjærstyrke": kdata.get("remould", [None]*len(depths))[i] if i < len(kdata.get("remould", [])) else None,
                "Uforstyrret skjærstyrke konus": kdata.get("undist", [None]*len(depths))[i] if i < len(kdata.get("undist", [])) else None,
                "Sensitivitet": kdata.get("sensitivity", [None]*len(depths))[i] if i < len(kdata.get("sensitivity", [])) else None,
                "Skjærstyrke enaks": edata.get("strength", [None]*len(depths))[i] if i < len(edata.get("strength", [])) else None,
                "Bruddtøyning": edata.get("deform", [None]*len(depths))[i] if i < len(edata.get("deform", [])) else None,
            }
            rows.append(row)

    df = pd.DataFrame(rows)
    df.to_excel(outfile_xlsx, index=False)
    print(f"✅ Excel table exported: {outfile_xlsx}")
    return outfile_xlsx
