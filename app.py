import os
import tempfile
import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Import your export functions
from plot_pdf import export_curfc_pdf, export_cu_enaks_konus_pdf

# -----------------------
# Logo path (baked in from repo)
# -----------------------
LOGO_PATH = os.path.join(os.path.dirname(__file__), "geovitalogo.png")

# -----------------------
# Streamlit UI
# -----------------------
st.title("Plot av udrenert skjærstyrke")

# --- Inputs ---
konus_files = st.file_uploader(
    "Last opp Konus-filer (Excel)", type=["xlsx", "xls", "xlsm"], accept_multiple_files=True
)

enaks_files = st.file_uploader(
    "Last opp Enaks-filer (Excel)", type=["xlsx", "xls", "xlsm"], accept_multiple_files=True
)

terrain_file = st.file_uploader("Last opp terrengnivåer (Excel)", type=["xlsx"])

sheet_name = st.text_input("Sheet-navn (kommaseparert hvis flere)", "Sheet 001")

# Ranges (fixed by type)
depth_range = "F6:F30"
konus_omrort_range = "M6:M30"
konus_uforstyrret_range = "L6:L30"
enaks_range = "G6:G30"

# Title block
st.subheader("Tittelblokk")
rapport_nr = st.text_input("Rapport Nr.", "XX")
dato       = st.date_input("Dato")
tegn       = st.text_input("Tegn", "IGH")
kontr      = st.text_input("Kontr", "JOG")
godkj      = st.text_input("Godkj", "AGR")
figur_nr   = st.text_input("Figur Nr.", "C3")

# -----------------------
# Helper: extract series
# -----------------------
def extract_series(files, sheet_names, x_range, y_range, terrain_lookup):
    series = []
    for file in files:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(file.getbuffer())
            tmp_path = tmp.name
        try:
            wb = load_workbook(tmp_path, data_only=True)
            borehole_name = None
            all_x, all_y = [], []
            for sname in sheet_names:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]
                if borehole_name is None:
                    borehole_name = ws["B6"].value
                x_vals = [cell[0].value for cell in ws[x_range]]
                y_vals = [cell[0].value for cell in ws[y_range]]
                pts = [(x, d) for x, d in zip(x_vals, y_vals) if x is not None and d is not None]
                if pts:
                    xs, ys = zip(*pts)
                    all_x.extend(xs)
                    all_y.extend(ys)
            if not borehole_name or not all_x:
                continue
            bh_key = str(borehole_name).strip().lower()
            Z = terrain_lookup.get(bh_key)
            if Z is None:
                continue
            elevs = [Z - d for d in all_y]
            # Sort by depth
            sorted_pts = sorted(zip(all_x, all_y, elevs), key=lambda t: t[1])
            xs, ys, elevs_sorted = zip(*sorted_pts)
            series.append((str(borehole_name), xs, ys, elevs_sorted, Z))
        except Exception as e:
            st.error(f"Feil ved lesing av {file.name}: {e}")
        finally:
            os.unlink(tmp_path)
    return series

# -----------------------
# Process when ready
# -----------------------
if st.button("Generer rapport"):
    if not konus_files or not terrain_file:
        st.error("Du må laste opp minst én Konus-fil og terrengnivåfil")
    else:
        with tempfile.TemporaryDirectory() as tmpdir:
            # Terrain
            terrain_path = os.path.join(tmpdir, terrain_file.name)
            with open(terrain_path, "wb") as f:
                f.write(terrain_file.getbuffer())
            terrain_df = pd.read_excel(terrain_path, usecols="A:B")
            terrain_df.columns = ["BH", "Z"]
            terrain_lookup = {
                str(bh).strip().lower(): z
                for bh, z in zip(terrain_df["BH"], terrain_df["Z"])
                if pd.notna(bh) and pd.notna(z)
            }

            sheet_names = [s.strip() for s in sheet_name.split(",") if s.strip()]

            # --- Konus Omrørt (alone) ---
            series_konus_omrort = extract_series(konus_files, sheet_names,
                                                 konus_omrort_range, depth_range, terrain_lookup)

            output_pdf_omrort = os.path.join(tmpdir, "konus_omrort.pdf")
            output_png_omrort = os.path.join(tmpdir, "konus_omrort.png")

            export_curfc_pdf(
                series_konus_omrort,
                outfile_pdf=output_pdf_omrort,
                outfile_png=output_png_omrort,
                logo_path=LOGO_PATH,
                title_info={
                    "rapport_nr": rapport_nr,
                    "dato": str(dato),
                    "tegn": tegn,
                    "kontr": kontr,
                    "godkj": godkj,
                    "figur_nr": figur_nr,
                },
            )

            # --- Konus Uforstyrret + Enaks ---
            series_konus_uforstyrret = extract_series(konus_files, sheet_names,
                                                      konus_uforstyrret_range, depth_range, terrain_lookup)
            series_enaks = []
            if enaks_files:
                series_enaks = extract_series(enaks_files, sheet_names,
                                              enaks_range, depth_range, terrain_lookup)

            output_pdf_uforstyrret = os.path.join(tmpdir, "konus_uforstyrret_enaks.pdf")
            output_png_uforstyrret = os.path.join(tmpdir, "konus_uforstyrret_enaks.png")

            export_cu_enaks_konus_pdf(
                series_konus_uforstyrret,
                series_enaks,
                outfile_pdf=output_pdf_uforstyrret,
                outfile_png=output_png_uforstyrret,
                logo_path=LOGO_PATH,
                title_info={
                    "rapport_nr": rapport_nr,
                    "dato": str(dato),
                    "tegn": tegn,
                    "kontr": kontr,
                    "godkj": godkj,
                    "figur_nr": figur_nr,
                },
            )

            # --- Show previews and download links ---
            st.subheader("Konus Omrørt (alene)")
            st.image(output_png_omrort, caption="Forhåndsvisning Konus Omrørt", use_column_width=True)
            with open(output_pdf_omrort, "rb") as f:
                st.download_button("Last ned PDF (Konus Omrørt)", f, file_name="konus_omrort.pdf")
            with open(output_png_omrort, "rb") as f:
                st.download_button("Last ned PNG (Konus Omrørt)", f, file_name="konus_omrort.png")

            st.subheader("Konus Uforstyrret + Enaks")
            st.image(output_png_uforstyrret, caption="Forhåndsvisning Konus Uforstyrret + Enaks", use_column_width=True)
            with open(output_pdf_uforstyrret, "rb") as f:
                st.download_button("Last ned PDF (Konus Uforstyrret + Enaks)", f, file_name="konus_uforstyrret_enaks.pdf")
            with open(output_png_uforstyrret, "rb") as f:
                st.download_button("Last ned PNG (Konus Uforstyrret + Enaks)", f, file_name="konus_uforstyrret_enaks.png")
